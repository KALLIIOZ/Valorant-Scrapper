import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def agregar_separador_fechas(df):
    """Inserta una fila en blanco entre grupos de fechas diferentes"""
    if df.empty or 'Fecha' not in df.columns:
        return df
    
    filas_insertadas = []
    fecha_anterior = None
    
    for idx, row in df.iterrows():
        fecha_actual = row['Fecha']
        
        # Si la fecha cambió y no es la primera iteración, agregar fila en blanco
        if fecha_anterior is not None and fecha_anterior != fecha_actual:
            fila_blanca = pd.Series({col: '' for col in df.columns})
            filas_insertadas.append(fila_blanca)
        
        filas_insertadas.append(row)
        fecha_anterior = fecha_actual
    
    return pd.DataFrame(filas_insertadas).reset_index(drop=True)


def exportar_stats_a_xlsx(carpeta='stats', archivo_salida='AX_Stats.xlsx', carpeta_premier='stats_premier'):
    """
    Lee todos los archivos JSON de las carpetas de stats y los exporta a un archivo XLSX.
    Los datos se organizan por fecha sin sobrescribir información anterior.
    Incluye comparación de rendimiento con respecto a la fecha anterior.
    Crea dos hojas: una para Competitivo y otra para Premier.
    
    Args:
        carpeta (str): Ruta de la carpeta donde están los JSON de Competitivo
        archivo_salida (str): Nombre del archivo XLSX de salida
        carpeta_premier (str): Ruta de la carpeta donde están los JSON de Premier
    """
    
    if not os.path.exists(carpeta):
        print(f"Error: La carpeta {carpeta} no existe")
        return False
    
    # Leer datos de Competitivo
    datos_competitivo = leer_stats_carpeta(carpeta)
    if not datos_competitivo:
        print("No se encontraron archivos JSON en stats")
        return False
    
    # Leer datos de Premier
    datos_premier = leer_stats_carpeta(carpeta_premier) if os.path.exists(carpeta_premier) else None
    
    fecha_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Procesar Competitivo
    df_competitivo_nuevo = pd.DataFrame(datos_competitivo)
    df_competitivo_nuevo['Fecha'] = fecha_actual
    
    # Remover columna SeasonIdPremier de Competitivo
    if 'SeasonIdPremier' in df_competitivo_nuevo.columns:
        df_competitivo_nuevo = df_competitivo_nuevo.drop('SeasonIdPremier', axis=1)
    
    columnas_numericas_comp = [col for col in df_competitivo_nuevo.columns 
                               if col not in ['Jugador', 'Fecha', 'Acto'] and col != 'Comparación']
    
    # Procesar Premier
    if datos_premier:
        df_premier_nuevo = pd.DataFrame(datos_premier)
        df_premier_nuevo['Fecha'] = fecha_actual
        columnas_numericas_prem = [col for col in df_premier_nuevo.columns 
                                   if col not in ['Jugador', 'Fecha', 'Season', 'SeasonId'] and col != 'Comparación']
    
    # Manejo de archivo existente
    if os.path.exists(archivo_salida):
        try:
            df_competitivo_existente = pd.read_excel(archivo_salida, sheet_name='Competitivo')
            # Remover SeasonIdPremier si existe
            if 'SeasonIdPremier' in df_competitivo_existente.columns:
                df_competitivo_existente = df_competitivo_existente.drop('SeasonIdPremier', axis=1)
            df_competitivo = agregar_cambios(df_competitivo_nuevo, df_competitivo_existente, columnas_numericas_comp)
            df_competitivo = pd.concat([df_competitivo_existente, df_competitivo], ignore_index=True)
            df_competitivo = agregar_separador_fechas(df_competitivo)
        except Exception as e:
            print(f"Error al leer Competitivo existente: {e}")
            df_competitivo = df_competitivo_nuevo
        
        if datos_premier:
            try:
                # Verificar si la hoja Premier existe
                xl_file = pd.ExcelFile(archivo_salida)
                if 'Premier' in xl_file.sheet_names:
                    df_premier_existente = pd.read_excel(archivo_salida, sheet_name='Premier')
                    df_premier = agregar_cambios(df_premier_nuevo, df_premier_existente, columnas_numericas_prem) # type: ignore
                    df_premier = pd.concat([df_premier_existente, df_premier], ignore_index=True)
                    df_premier = agregar_separador_fechas(df_premier)
                else:
                    df_premier = df_premier_nuevo # type: ignore
            except Exception as e:
                print(f"Error al leer Premier existente: {e}")
                df_premier = df_premier_nuevo # type: ignore
    else:
        df_competitivo = df_competitivo_nuevo
        df_premier = df_premier_nuevo if datos_premier else None # type: ignore
    
    # Exportar a XLSX
    try:
        with pd.ExcelWriter(archivo_salida, engine='openpyxl', mode='w') as writer:
            df_competitivo.to_excel(writer, index=False, sheet_name='Competitivo')
            if df_premier is not None: # type: ignore
                df_premier.to_excel(writer, index=False, sheet_name='Premier') # type: ignore
        
        aplicar_formato_condicional(archivo_salida, df_competitivo, columnas_numericas_comp, 'Competitivo')
        if df_premier is not None: # type: ignore
            aplicar_formato_condicional(archivo_salida, df_premier, columnas_numericas_prem, 'Premier') # type: ignore
        
        print(f"Estadísticas exportadas a {archivo_salida}")
        return True
    except Exception as e:
        print(f"Error al exportar a XLSX: {e}")
        return False


def leer_stats_carpeta(carpeta):
    """Lee todos los JSON de una carpeta"""
    datos = []
    if not os.path.exists(carpeta):
        return datos
    
    for archivo in os.listdir(carpeta):
        if archivo.endswith('.json'):
            ruta_archivo = os.path.join(carpeta, archivo)
            try:
                with open(ruta_archivo, 'r') as f:
                    contenido = json.load(f)
                    datos.append(contenido)
            except Exception as e:
                print(f"Error al leer {archivo}: {e}")
    
    return datos


def agregar_cambios(df_nuevo, df_existente, columnas_numericas):
    """Agrega columnas de cambio comparando con el registro anterior"""
    for col in columnas_numericas:
        df_nuevo[f'{col}_Cambio'] = ''
        
        for idx, row in df_nuevo.iterrows():
            jugador = row['Jugador']
            valor_nuevo = row[col]
            
            registros_anteriores = df_existente[df_existente['Jugador'] == jugador]
            
            if not registros_anteriores.empty:
                valor_anterior = registros_anteriores.iloc[-1][col]
                
                try:
                    valor_nuevo_float = float(str(valor_nuevo).replace('%', ''))
                    valor_anterior_float = float(str(valor_anterior).replace('%', ''))
                    
                    diferencia = valor_nuevo_float - valor_anterior_float
                    
                    if diferencia > 0:
                        df_nuevo.at[idx, f'{col}_Cambio'] = f'+{diferencia:.2f}'
                    elif diferencia < 0:
                        df_nuevo.at[idx, f'{col}_Cambio'] = f'{diferencia:.2f}'
                    else:
                        df_nuevo.at[idx, f'{col}_Cambio'] = '0'
                except:
                    df_nuevo.at[idx, f'{col}_Cambio'] = 'N/A'
    
    return df_nuevo


def aplicar_formato_condicional(archivo_salida, df, columnas_numericas, nombre_hoja='Competitivo'):
    """
    Aplica formato condicional de colores a las columnas de cambio.
    Rojo para bajón de rendimiento, verde para mejora.
    También embellece el Excel con colores atractivos.
    """
    try:
        wb = load_workbook(archivo_salida)
        ws = wb[nombre_hoja]
        
        # Definir colores
        color_header = '4472C4'  # Azul oscuro para encabezados
        color_jugador = 'FFE699'  # Naranja claro
        color_fecha = 'C6EFCE'    # Verde claro
        color_columnas = 'D9E8F5' # Azul medio claro
        
        # Colorear encabezados y datos según columna
        for col_idx, col_name in enumerate(ws[1], 1): # type: ignore
            columna_actual = col_name.value
            
            # Colorear encabezado
            celda_header = ws.cell(row=1, column=col_idx) # type: ignore
            celda_header.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
            celda_header.font = Font(bold=True, color='FFFFFF', size=11)
            celda_header.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorear datos según tipo de columna
            for row_idx in range(2, ws.max_row + 1): # type: ignore
                celda = ws.cell(row=row_idx, column=col_idx) # type: ignore
                
                if columna_actual == 'Jugador':
                    # Naranja claro para nombres
                    celda.fill = PatternFill(start_color=color_jugador, end_color=color_jugador, fill_type='solid')
                    celda.font = Font(bold=True, size=10)
                elif columna_actual == 'Fecha':
                    # Verde claro para fechas
                    celda.fill = PatternFill(start_color=color_fecha, end_color=color_fecha, fill_type='solid')
                    celda.font = Font(size=10)
                elif '_Cambio' in str(columna_actual):
                    # Ya tiene formato condicional de colores rojo/verde
                    valor = celda.value
                    if valor and valor != 'N/A':
                        try:
                            numero = float(str(valor).replace('+', ''))
                            
                            if numero > 0:
                                # Verde para mejora
                                celda.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                                celda.font = Font(bold=True, color='FFFFFF')
                            elif numero < 0:
                                # Rojo para bajón
                                celda.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                                celda.font = Font(bold=True, color='FFFFFF')
                            else:
                                # Gris para sin cambios
                                celda.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                                celda.font = Font(bold=True)
                            
                            celda.alignment = Alignment(horizontal='center', vertical='center')
                        except:
                            pass
                else:
                    # Azul medio claro para otras columnas
                    celda.fill = PatternFill(start_color=color_columnas, end_color=color_columnas, fill_type='solid')
                    celda.font = Font(size=10)
                
                # Alinear al centro si no es la columna de nombre
                if columna_actual != 'Jugador':
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    celda.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho de columnas
        for column in ws.columns: # type: ignore
            max_length = 0
            column_letter = column[0].column_letter # type: ignore
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width # type: ignore
        
        # Congelar encabezados
        ws.freeze_panes = 'A2' # type: ignore
        
        wb.save(archivo_salida)
        print("Formato embellecido aplicado correctamente")
    except Exception as e:
        print(f"Error al aplicar formato: {e}")
